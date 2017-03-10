#!/usr/bin/env python3
# -*- encoding: utf-8 -*-

import openpyxl



def test_openpyxl_read():
    wb = openpyxl.load_workbook('example.xlsx')
    print(type(wb))
    print(wb.get_sheet_names())

    sheet = wb.get_sheet_by_name('Sheet1')
    print(sheet['A1'].value)

    for i in range(1, 8):
        print(i, sheet.cell(row=i, column=2).value)

def test_openpyxl_write():
    wb = openpyxl.Workbook()
    sheet = wb.get_sheet_names()
    print(sheet)
    sheet = wb.get_active_sheet()
    print(sheet.title)
    sheet.title = 'Spam Bacon Eggs Sheet'
    print(wb.get_sheet_names())

    wb = openpyxl.load_workbook('example.xlsx')
    sheet = wb.get_active_sheet()
    sheet.title = 'Spam Spam Spam'
    wb.save('example_copy.xlsx')

if __name__ == '__main__':
    test_openpyxl_write()
